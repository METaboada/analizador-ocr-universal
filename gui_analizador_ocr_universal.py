import tkinter as tk
from tkinter import ttk, filedialog, messagebox, scrolledtext
import threading
import os
import subprocess
import re
from datetime import datetime
import json

class UniversalOCRAnalyzer:
    def __init__(self, root):
        self.root = root
        self.root.title("📄 Analizador OCR Universal - Búsqueda por Palabras Clave")
        self.root.geometry("900x700")
        self.root.minsize(800, 600)
        
        # Variables
        self.selected_files = []
        self.output_folder = tk.StringVar()
        self.is_processing = False
        
        # Verificar dependencias al iniciar
        self.verificar_dependencias_inicial()
        self.search_keywords = []
        self.project_name = tk.StringVar(value="Proyecto_OCR")
        
        # Configurar estilo
        self.setup_style()
        
        # Crear interfaz
        self.create_widgets()
        
        # Configurar output folder por defecto
        self.output_folder.set(os.path.join(os.getcwd(), "resultados_ocr"))
        
        # Cargar palabras clave por defecto
        self.load_default_keywords()

    def setup_style(self):
        """Configurar el estilo de la interfaz"""
        style = ttk.Style()
        style.theme_use('clam')
        
        # Configurar colores
        style.configure('Title.TLabel', font=('Arial', 16, 'bold'), foreground='#2c3e50')
        style.configure('Subtitle.TLabel', font=('Arial', 10, 'bold'), foreground='#34495e')
        style.configure('Success.TLabel', foreground='#27ae60')
        style.configure('Error.TLabel', foreground='#e74c3c')

    def create_widgets(self):
        """Crear todos los widgets de la interfaz"""
        
        # Frame principal con padding
        main_frame = ttk.Frame(self.root, padding="20")
        main_frame.grid(row=0, column=0, sticky="nsew")
        
        # Configurar grid weights
        self.root.columnconfigure(0, weight=1)
        self.root.rowconfigure(0, weight=1)
        main_frame.columnconfigure(0, weight=1)
        
        # Título
        title_label = ttk.Label(main_frame, text="📄 Analizador OCR Universal", style='Title.TLabel')
        title_label.grid(row=0, column=0, pady=(0, 20), sticky="ew")
        
        # Crear notebook para pestañas
        self.notebook = ttk.Notebook(main_frame)
        self.notebook.grid(row=1, column=0, sticky="nsew", pady=(0, 10))
        main_frame.rowconfigure(1, weight=1)
        
        # Pestaña 1: Configuración del proyecto
        self.create_project_tab()
        
        # Pestaña 2: Selección de archivos
        self.create_files_tab()
        
        # Pestaña 3: Palabras clave
        self.create_keywords_tab()
        
        # Pestaña 4: Procesamiento
        self.create_processing_tab()
        
        # Frame de botones principales
        buttons_frame = ttk.Frame(main_frame)
        buttons_frame.grid(row=2, column=0, sticky="ew", pady=(10, 0))
        buttons_frame.columnconfigure(1, weight=1)
        
        # Botón principal de análisis
        self.analyze_btn = ttk.Button(buttons_frame, text="🚀 Iniciar Análisis OCR", 
                                     command=self.start_analysis, width=20)
        self.analyze_btn.grid(row=0, column=0, padx=(0, 10))
        
        # Barra de progreso
        self.progress_bar = ttk.Progressbar(buttons_frame, mode='indeterminate')
        self.progress_bar.grid(row=0, column=1, sticky="ew", padx=(10, 10))
        
        # Botón de ayuda
        help_btn = ttk.Button(buttons_frame, text="❓ Ayuda", 
                             command=self.show_help, width=10)
        help_btn.grid(row=0, column=2)

    def create_project_tab(self):
        """Crear pestaña de configuración del proyecto"""
        project_frame = ttk.Frame(self.notebook, padding="20")
        self.notebook.add(project_frame, text="🏷️ Proyecto")
        
        project_frame.columnconfigure(1, weight=1)
        
        # Nombre del proyecto
        ttk.Label(project_frame, text="Nombre del proyecto:", font=('Arial', 10, 'bold')).grid(
            row=0, column=0, sticky="w", pady=(0, 10), padx=(0, 10))
        
        project_entry = ttk.Entry(project_frame, textvariable=self.project_name, font=('Arial', 10))
        project_entry.grid(row=0, column=1, sticky="ew", pady=(0, 10))
        
        # Descripción del proyecto
        ttk.Label(project_frame, text="Descripción:", font=('Arial', 10, 'bold')).grid(
            row=1, column=0, sticky="nw", pady=(0, 10), padx=(0, 10))
        
        self.project_description = scrolledtext.ScrolledText(project_frame, height=4, width=50)
        self.project_description.grid(row=1, column=1, sticky="ew", pady=(0, 10))
        self.project_description.insert("1.0", "Análisis OCR de documentos PDF para búsqueda de palabras clave específicas.")
        
        # Carpeta de salida
        ttk.Label(project_frame, text="Carpeta de resultados:", font=('Arial', 10, 'bold')).grid(
            row=2, column=0, sticky="w", pady=(0, 10), padx=(0, 10))
        
        output_frame = ttk.Frame(project_frame)
        output_frame.grid(row=2, column=1, sticky="ew", pady=(0, 10))
        output_frame.columnconfigure(0, weight=1)
        
        output_entry = ttk.Entry(output_frame, textvariable=self.output_folder, font=('Arial', 9))
        output_entry.grid(row=0, column=0, sticky="ew", padx=(0, 10))
        
        browse_btn = ttk.Button(output_frame, text="📁 Examinar", 
                               command=self.browse_output_folder, width=12)
        browse_btn.grid(row=0, column=1)

    def create_files_tab(self):
        """Crear pestaña de selección de archivos"""
        files_frame = ttk.Frame(self.notebook, padding="20")
        self.notebook.add(files_frame, text="📁 Archivos")
        
        files_frame.columnconfigure(0, weight=1)
        files_frame.rowconfigure(2, weight=1)
        
        # Botones de archivos
        files_buttons_frame = ttk.Frame(files_frame)
        files_buttons_frame.grid(row=0, column=0, sticky="ew", pady=(0, 10))
        
        select_files_btn = ttk.Button(files_buttons_frame, text="📁 Seleccionar PDFs", 
                                     command=self.select_files, width=20)
        select_files_btn.pack(side=tk.LEFT, padx=(0, 10))
        
        remove_files_btn = ttk.Button(files_buttons_frame, text="🗑️ Quitar Seleccionados", 
                                     command=self.remove_selected_files, width=20)
        remove_files_btn.pack(side=tk.LEFT, padx=(0, 10))
        
        clear_all_btn = ttk.Button(files_buttons_frame, text="🧹 Limpiar Todo", 
                                  command=self.clear_all_files, width=15)
        clear_all_btn.pack(side=tk.LEFT)
        
        # Label con contador de archivos
        self.files_count_label = ttk.Label(files_frame, text="No hay archivos seleccionados", 
                                          font=('Arial', 10, 'bold'))
        self.files_count_label.grid(row=1, column=0, sticky="w", pady=(5, 10))
        
        # Lista de archivos seleccionados con scrollbar
        files_list_frame = ttk.Frame(files_frame)
        files_list_frame.grid(row=2, column=0, sticky="nsew")
        files_list_frame.columnconfigure(0, weight=1)
        files_list_frame.rowconfigure(0, weight=1)
        
        self.files_listbox = tk.Listbox(files_list_frame, selectmode=tk.EXTENDED, font=('Arial', 9))
        self.files_listbox.grid(row=0, column=0, sticky="nsew")
        
        files_scrollbar = ttk.Scrollbar(files_list_frame, orient="vertical", command=self.files_listbox.yview)
        files_scrollbar.grid(row=0, column=1, sticky="ns")
        self.files_listbox.configure(yscrollcommand=files_scrollbar.set)

    def create_keywords_tab(self):
        """Crear pestaña de palabras clave"""
        keywords_frame = ttk.Frame(self.notebook, padding="20")
        self.notebook.add(keywords_frame, text="🔍 Palabras Clave")
        
        keywords_frame.columnconfigure(0, weight=1)
        keywords_frame.rowconfigure(1, weight=1)
        
        # Frame superior con botones
        keywords_top_frame = ttk.Frame(keywords_frame)
        keywords_top_frame.grid(row=0, column=0, sticky="ew", pady=(0, 10))
        keywords_top_frame.columnconfigure(2, weight=1)
        
        # Entrada de nueva palabra clave
        ttk.Label(keywords_top_frame, text="Nueva palabra clave:", font=('Arial', 10, 'bold')).grid(
            row=0, column=0, sticky="w", padx=(0, 10))
        
        self.keyword_entry = ttk.Entry(keywords_top_frame, font=('Arial', 10))
        self.keyword_entry.grid(row=0, column=1, sticky="ew", padx=(0, 10))
        self.keyword_entry.bind('<Return>', lambda e: self.add_keyword())
        
        add_keyword_btn = ttk.Button(keywords_top_frame, text="➕ Agregar", 
                                    command=self.add_keyword, width=10)
        add_keyword_btn.grid(row=0, column=2, padx=(0, 10))
        
        # Botones de gestión
        keywords_buttons_frame = ttk.Frame(keywords_frame)
        keywords_buttons_frame.grid(row=1, column=0, sticky="ew", pady=(0, 10))
        
        remove_keyword_btn = ttk.Button(keywords_buttons_frame, text="🗑️ Quitar Seleccionada", 
                                       command=self.remove_keyword, width=18)
        remove_keyword_btn.pack(side=tk.LEFT, padx=(0, 10))
        
        clear_keywords_btn = ttk.Button(keywords_buttons_frame, text="🧹 Limpiar Todas", 
                                       command=self.clear_keywords, width=15)
        clear_keywords_btn.pack(side=tk.LEFT, padx=(0, 10))
        
        load_keywords_btn = ttk.Button(keywords_buttons_frame, text="📂 Cargar desde archivo", 
                                      command=self.load_keywords_file, width=18)
        load_keywords_btn.pack(side=tk.LEFT, padx=(0, 10))
        
        save_keywords_btn = ttk.Button(keywords_buttons_frame, text="💾 Guardar en archivo", 
                                      command=self.save_keywords_file, width=18)
        save_keywords_btn.pack(side=tk.LEFT)
        
        # Lista de palabras clave
        keywords_list_frame = ttk.Frame(keywords_frame)
        keywords_list_frame.grid(row=2, column=0, sticky="nsew", pady=(10, 0))
        keywords_list_frame.columnconfigure(0, weight=1)
        keywords_list_frame.rowconfigure(0, weight=1)
        
        self.keywords_listbox = tk.Listbox(keywords_list_frame, selectmode=tk.EXTENDED, font=('Arial', 10))
        self.keywords_listbox.grid(row=0, column=0, sticky="nsew")
        
        keywords_scrollbar = ttk.Scrollbar(keywords_list_frame, orient="vertical", 
                                          command=self.keywords_listbox.yview)
        keywords_scrollbar.grid(row=0, column=1, sticky="ns")
        self.keywords_listbox.configure(yscrollcommand=keywords_scrollbar.set)
        
        # Contador de palabras clave
        self.keywords_count_label = ttk.Label(keywords_frame, text="0 palabras clave configuradas", 
                                             font=('Arial', 10, 'bold'))
        self.keywords_count_label.grid(row=3, column=0, sticky="w", pady=(10, 0))

    def create_processing_tab(self):
        """Crear pestaña de procesamiento y resultados"""
        processing_frame = ttk.Frame(self.notebook, padding="20")
        self.notebook.add(processing_frame, text="⚙️ Procesamiento")
        
        processing_frame.columnconfigure(0, weight=1)
        processing_frame.rowconfigure(1, weight=1)
        
        # Frame de configuración
        config_frame = ttk.LabelFrame(processing_frame, text="Configuración del Análisis", padding="10")
        config_frame.grid(row=0, column=0, sticky="ew", pady=(0, 10))
        config_frame.columnconfigure(1, weight=1)
        
        # Opciones de procesamiento
        self.ocr_enabled = tk.BooleanVar(value=True)
        ocr_check = ttk.Checkbutton(config_frame, text="Usar OCR para documentos escaneados", 
                                   variable=self.ocr_enabled)
        ocr_check.grid(row=0, column=0, columnspan=2, sticky="w", pady=(0, 5))
        
        self.case_sensitive = tk.BooleanVar(value=False)
        case_check = ttk.Checkbutton(config_frame, text="Búsqueda sensible a mayúsculas/minúsculas", 
                                    variable=self.case_sensitive)
        case_check.grid(row=1, column=0, columnspan=2, sticky="w", pady=(0, 5))
        
        self.whole_words = tk.BooleanVar(value=True)
        words_check = ttk.Checkbutton(config_frame, text="Solo palabras completas", 
                                     variable=self.whole_words)
        words_check.grid(row=2, column=0, columnspan=2, sticky="w")
        
        # Log de procesamiento
        log_frame = ttk.LabelFrame(processing_frame, text="Log de Procesamiento", padding="10")
        log_frame.grid(row=1, column=0, sticky="nsew")
        log_frame.columnconfigure(0, weight=1)
        log_frame.rowconfigure(0, weight=1)
        
        self.log_text = scrolledtext.ScrolledText(log_frame, height=15, font=('Consolas', 9))
        self.log_text.grid(row=0, column=0, sticky="nsew")

    def load_default_keywords(self):
        """Cargar palabras clave por defecto"""
        default_keywords = [
            "administración", "gobierno", "municipal", "público", "servicio",
            "ciudad", "urbano", "infraestructura", "desarrollo", "gestión"
        ]
        
        self.search_keywords = default_keywords.copy()
        self.update_keywords_display()

    def add_keyword(self):
        """Agregar una nueva palabra clave"""
        keyword = self.keyword_entry.get().strip()
        if keyword and keyword not in self.search_keywords:
            self.search_keywords.append(keyword)
            self.update_keywords_display()
            self.keyword_entry.delete(0, tk.END)
        elif keyword in self.search_keywords:
            messagebox.showwarning("Palabra duplicada", f"La palabra '{keyword}' ya está en la lista.")

    def remove_keyword(self):
        """Remover palabra clave seleccionada"""
        selection = self.keywords_listbox.curselection()
        if selection:
            for index in reversed(selection):
                del self.search_keywords[index]
            self.update_keywords_display()

    def clear_keywords(self):
        """Limpiar todas las palabras clave"""
        if messagebox.askyesno("Confirmar", "¿Está seguro de que desea eliminar todas las palabras clave?"):
            self.search_keywords.clear()
            self.update_keywords_display()

    def update_keywords_display(self):
        """Actualizar la visualización de palabras clave"""
        self.keywords_listbox.delete(0, tk.END)
        for keyword in self.search_keywords:
            self.keywords_listbox.insert(tk.END, keyword)
        
        count = len(self.search_keywords)
        self.keywords_count_label.config(text=f"{count} palabra{'s' if count != 1 else ''} clave configurada{'s' if count != 1 else ''}")

    def load_keywords_file(self):
        """Cargar palabras clave desde archivo"""
        file_path = filedialog.askopenfilename(
            title="Cargar palabras clave",
            filetypes=[("Archivos de texto", "*.txt"), ("Archivos JSON", "*.json"), ("Todos los archivos", "*.*")],
            initialfile="palabras_clave.txt"
        )
        
        if file_path:
            try:
                with open(file_path, 'r', encoding='utf-8') as file:
                    if file_path.endswith('.json'):
                        # Soporte para JSON (opcional)
                        data = json.load(file)
                        if isinstance(data, list):
                            keywords = data
                        elif isinstance(data, dict) and 'keywords' in data:
                            keywords = data['keywords']
                        else:
                            raise ValueError("Formato JSON no válido")
                    else:
                        # Archivo de texto, una palabra por línea (formato principal)
                        # Ignora líneas vacías y comentarios que empiecen con #
                        keywords = []
                        for line in file.readlines():
                            line = line.strip()
                            if line and not line.startswith('#'):
                                keywords.append(line)
                
                # Agregar palabras únicas
                added_count = 0
                for keyword in keywords:
                    if keyword not in self.search_keywords:
                        self.search_keywords.append(keyword)
                        added_count += 1
                
                self.update_keywords_display()
                messagebox.showinfo("Éxito", f"Se agregaron {added_count} palabras clave nuevas.")
                
            except Exception as e:
                messagebox.showerror("Error", f"No se pudo cargar el archivo:\n{str(e)}")

    def save_keywords_file(self):
        """Guardar palabras clave en archivo"""
        if not self.search_keywords:
            messagebox.showwarning("Sin datos", "No hay palabras clave para guardar.")
            return
        
        file_path = filedialog.asksaveasfilename(
            title="Guardar palabras clave",
            defaultextension=".txt",
            initialfile=f"palabras_clave_{self.project_name.get().replace(' ', '_')}.txt",
            filetypes=[("Archivos de texto", "*.txt"), ("Archivos JSON", "*.json")]
        )
        
        if file_path:
            try:
                with open(file_path, 'w', encoding='utf-8') as file:
                    if file_path.endswith('.json'):
                        # Opción JSON (para casos especiales)
                        json.dump({
                            "proyecto": self.project_name.get(),
                            "keywords": self.search_keywords,
                            "fecha_creacion": datetime.now().isoformat()
                        }, file, indent=2, ensure_ascii=False)
                    else:
                        # Formato TXT (predeterminado y recomendado)
                        file.write(f"# Palabras clave para: {self.project_name.get()}\n")
                        file.write(f"# Creado: {datetime.now().strftime('%d/%m/%Y %H:%M:%S')}\n")
                        file.write(f"# Total de palabras: {len(self.search_keywords)}\n")
                        file.write("#\n")
                        file.write("# Una palabra clave por línea:\n")
                        file.write("#" + "-" * 40 + "\n")
                        for keyword in self.search_keywords:
                            file.write(keyword + '\n')
                
                messagebox.showinfo("Éxito", f"Palabras clave guardadas en:\n{file_path}")
                
            except Exception as e:
                messagebox.showerror("Error", f"No se pudo guardar el archivo:\n{str(e)}")

    def select_files(self):
        """Seleccionar archivos PDF"""
        files = filedialog.askopenfilenames(
            title="Seleccionar archivos PDF",
            filetypes=[("Archivos PDF", "*.pdf"), ("Todos los archivos", "*.*")]
        )
        
        for file in files:
            if file not in self.selected_files:
                self.selected_files.append(file)
        
        self.update_files_display()

    def remove_selected_files(self):
        """Remover archivos seleccionados"""
        selection = self.files_listbox.curselection()
        if selection:
            for index in reversed(selection):
                del self.selected_files[index]
            self.update_files_display()

    def clear_all_files(self):
        """Limpiar todos los archivos"""
        if messagebox.askyesno("Confirmar", "¿Está seguro de que desea eliminar todos los archivos?"):
            self.selected_files.clear()
            self.update_files_display()

    def update_files_display(self):
        """Actualizar la visualización de archivos"""
        self.files_listbox.delete(0, tk.END)
        for file in self.selected_files:
            self.files_listbox.insert(tk.END, os.path.basename(file))
        
        count = len(self.selected_files)
        if count == 0:
            self.files_count_label.config(text="No hay archivos seleccionados")
        else:
            self.files_count_label.config(text=f"{count} archivo{'s' if count != 1 else ''} seleccionado{'s' if count != 1 else ''}")

    def browse_output_folder(self):
        """Examinar carpeta de salida"""
        folder = filedialog.askdirectory(title="Seleccionar carpeta de resultados")
        if folder:
            self.output_folder.set(folder)

    def log_message(self, message):
        """Agregar mensaje al log"""
        timestamp = datetime.now().strftime("%H:%M:%S")
        self.log_text.insert(tk.END, f"[{timestamp}] {message}\n")
        self.log_text.see(tk.END)
        self.root.update_idletasks()

    def start_analysis(self):
        """Iniciar el análisis OCR"""
        if not self.selected_files:
            messagebox.showwarning("Sin archivos", "Por favor seleccione al menos un archivo PDF.")
            return
        
        if not self.search_keywords:
            messagebox.showwarning("Sin palabras clave", "Por favor configure al menos una palabra clave.")
            return
        
        if self.is_processing:
            messagebox.showinfo("Procesando", "Ya hay un análisis en progreso.")
            return
        
        # Confirmar inicio
        if not messagebox.askyesno("Confirmar análisis", 
                                  f"¿Iniciar análisis de {len(self.selected_files)} archivo(s) "
                                  f"con {len(self.search_keywords)} palabra(s) clave?"):
            return
        
        # Iniciar procesamiento en hilo separado
        self.is_processing = True
        self.analyze_btn.config(state='disabled', text="🔄 Procesando...")
        self.progress_bar.start()
        
        # Cambiar a pestaña de procesamiento
        self.notebook.select(3)
        
        # Limpiar log
        self.log_text.delete(1.0, tk.END)
        
        # Iniciar hilo de procesamiento
        thread = threading.Thread(target=self.run_analysis)
        thread.daemon = True
        thread.start()

    def run_analysis(self):
        """Ejecutar el análisis en hilo separado"""
        try:
            self.log_message("🚀 Iniciando análisis OCR universal...")
            self.log_message(f"📋 Proyecto: {self.project_name.get()}")
            self.log_message(f"📁 Archivos a procesar: {len(self.selected_files)}")
            self.log_message(f"🔍 Palabras clave: {len(self.search_keywords)}")
            self.log_message("")
            
            # Crear carpeta de salida
            output_dir = self.output_folder.get()
            os.makedirs(output_dir, exist_ok=True)
            
            # Importar librerías necesarias
            try:
                import PyPDF2
                import pytesseract
                from pdf2image import convert_from_path
                from PIL import Image
                
                # Configurar pytesseract y TESSDATA_PREFIX
                pytesseract.pytesseract.tesseract_cmd = r'C:\Program Files\Tesseract-OCR\tesseract.exe'
                
                # Configurar TESSDATA_PREFIX para usar archivos locales
                script_dir = os.path.dirname(os.path.abspath(__file__))
                tessdata_path = os.path.join(script_dir, 'tessdata')
                os.environ['TESSDATA_PREFIX'] = tessdata_path
                
                # Hacer disponibles para las funciones
                self.PyPDF2 = PyPDF2
                self.pytesseract = pytesseract
                self.convert_from_path = convert_from_path
                
            except ImportError as e:
                self.log_message(f"❌ Error al importar librerías: {str(e)}")
                self.log_message("Por favor, ejecute 'python verificar_instalacion.py'")
                return
            
            results = []
            
            for i, file_path in enumerate(self.selected_files, 1):
                filename = os.path.basename(file_path)
                self.log_message(f"🔄 Procesando {i}/{len(self.selected_files)}: {filename}")
                
                file_results = {
                    'file': filename,
                    'file_path': file_path,
                    'matches': [],
                    'total_matches': 0,
                    'pages_processed': 0,
                    'error': None
                }
                
                try:
                    # Procesar PDF
                    file_matches = self.process_pdf_file(file_path)
                    file_results['matches'] = file_matches
                    file_results['total_matches'] = len(file_matches)
                    file_results['pages_processed'] = max([m['page'] for m in file_matches]) if file_matches else 0
                    
                    self.log_message(f"✅ {filename}: {file_results['total_matches']} coincidencias en {file_results['pages_processed']} páginas")
                    
                except Exception as e:
                    error_msg = f"Error procesando {filename}: {str(e)}"
                    self.log_message(f"❌ {error_msg}")
                    file_results['error'] = error_msg
                
                results.append(file_results)
            
            # Generar reporte detallado
            report_path = self.generate_detailed_report(results, output_dir)
            
            self.log_message("")
            self.log_message("🎉 ¡Análisis completado exitosamente!")
            self.log_message(f"📊 Resultados guardados en: {output_dir}")
            self.log_message(f"📁 Ruta completa: {os.path.abspath(output_dir)}")
            
            # Mostrar mensaje de éxito con opción de abrir resultados
            self.root.after(0, lambda: self.show_completion_dialog(os.path.abspath(output_dir), report_path))
            
        except Exception as e:
            error_msg = f"❌ Error durante el análisis: {str(e)}"
            self.log_message(error_msg)
            self.root.after(0, lambda: messagebox.showerror("Error", error_msg))
        
        finally:
            # Restaurar interfaz
            self.root.after(0, self.analysis_finished)

    def process_pdf_file(self, file_path):
        """Procesar un archivo PDF y buscar palabras clave"""
        matches = []
        
        try:
            # Intentar leer texto directo del PDF primero
            with open(file_path, 'rb') as file:
                pdf_reader = self.PyPDF2.PdfReader(file)
                
                for page_num, page in enumerate(pdf_reader.pages, 1):
                    try:
                        # Extraer texto directo
                        text = page.extract_text()
                        
                        # Si no hay texto o es muy poco, usar OCR
                        if len(text.strip()) < 50:
                            self.log_message(f"   📝 Página {page_num}: Usando OCR (documento escaneado)")
                            text = self.extract_text_with_ocr(file_path, page_num)
                        else:
                            self.log_message(f"   📝 Página {page_num}: Texto extraído directamente")
                        
                        # Buscar palabras clave en el texto
                        page_matches = self.find_keywords_in_text(text, page_num)
                        matches.extend(page_matches)
                        
                    except Exception as e:
                        self.log_message(f"   ⚠️ Error en página {page_num}: {str(e)}")
                        
        except Exception as e:
            self.log_message(f"   ❌ Error leyendo PDF: {str(e)}")
            raise
        
        return matches

    def extract_text_with_ocr(self, file_path, page_num):
        """Extraer texto usando OCR para una página específica"""
        try:
            # Convertir página específica a imagen
            images = self.convert_from_path(file_path, first_page=page_num, last_page=page_num, dpi=300)
            
            if images:
                # Usar OCR en español
                text = self.pytesseract.image_to_string(images[0], lang='spa')
                return text
            else:
                return ""
                
        except Exception as e:
            self.log_message(f"   ⚠️ Error en OCR página {page_num}: {str(e)}")
            return ""

    def find_keywords_in_text(self, text, page_num):
        """Buscar palabras clave en el texto y devolver contexto"""
        matches = []
        
        if not text.strip():
            return matches
        
        # Configuración de búsqueda
        flags = 0 if self.case_sensitive.get() else re.IGNORECASE
        
        for keyword in self.search_keywords:
            # Crear patrón de búsqueda
            if self.whole_words.get():
                pattern = r'\b' + re.escape(keyword) + r'\b'
            else:
                pattern = re.escape(keyword)
            
            # Buscar todas las ocurrencias
            for match in re.finditer(pattern, text, flags):
                # Extraer contexto (50 caracteres antes y después)
                start_pos = max(0, match.start() - 50)
                end_pos = min(len(text), match.end() + 50)
                context = text[start_pos:end_pos].strip()
                
                # Limpiar contexto (remover saltos de línea excesivos)
                context = re.sub(r'\s+', ' ', context)
                
                # Resaltar la palabra encontrada en el contexto
                highlighted_context = context.replace(
                    match.group(), 
                    f"**{match.group()}**"
                )
                
                matches.append({
                    'page': page_num,
                    'keyword': keyword,
                    'matched_text': match.group(),
                    'context': context,
                    'highlighted_context': highlighted_context,
                    'position': match.start()
                })
        
        return matches

    def generate_detailed_report(self, results, output_dir):
        """Generar reporte detallado de resultados"""
        self.log_message("📊 Generando reporte detallado...")
        
        timestamp = datetime.now().strftime('%Y%m%d_%H%M%S')
        project_name = self.project_name.get().replace(' ', '_')
        
        # Generar reporte en texto
        report_path = os.path.join(output_dir, f"reporte_{project_name}_{timestamp}.txt")
        
        with open(report_path, 'w', encoding='utf-8') as f:
            f.write(f"REPORTE DETALLADO DE ANÁLISIS OCR - {self.project_name.get().upper()}\n")
            f.write("=" * 80 + "\n")
            f.write(f"Fecha: {datetime.now().strftime('%d/%m/%Y %H:%M:%S')}\n")
            f.write(f"Total de archivos procesados: {len(results)}\n")
            f.write(f"Palabras clave buscadas: {', '.join(self.search_keywords)}\n")
            f.write(f"Configuración:\n")
            f.write(f"  - Sensible a mayúsculas: {'Sí' if self.case_sensitive.get() else 'No'}\n")
            f.write(f"  - Solo palabras completas: {'Sí' if self.whole_words.get() else 'No'}\n")
            f.write(f"  - OCR habilitado: {'Sí' if self.ocr_enabled.get() else 'No'}\n")
            f.write("\n")
            
            # Estadísticas generales
            total_matches = sum(r['total_matches'] for r in results)
            files_with_matches = len([r for r in results if r['total_matches'] > 0])
            files_with_errors = len([r for r in results if r.get('error')])
            
            f.write("RESUMEN EJECUTIVO:\n")
            f.write("-" * 40 + "\n")
            f.write(f"• Total de coincidencias encontradas: {total_matches}\n")
            f.write(f"• Archivos con coincidencias: {files_with_matches} de {len(results)}\n")
            f.write(f"• Archivos con errores: {files_with_errors}\n")
            f.write(f"• Promedio de coincidencias por archivo: {total_matches/len(results):.1f}\n")
            f.write("\n")
            
            # Ranking de palabras clave más encontradas
            keyword_count = {}
            for result in results:
                for match in result.get('matches', []):
                    keyword = match['keyword']
                    keyword_count[keyword] = keyword_count.get(keyword, 0) + 1
            
            if keyword_count:
                f.write("PALABRAS CLAVE MÁS FRECUENTES:\n")
                f.write("-" * 40 + "\n")
                sorted_keywords = sorted(keyword_count.items(), key=lambda x: x[1], reverse=True)
                for i, (keyword, count) in enumerate(sorted_keywords[:10], 1):
                    f.write(f"{i:2d}. {keyword}: {count} ocurrencias\n")
                f.write("\n")
            
            # Detalle por archivo
            f.write("DETALLE POR ARCHIVO:\n")
            f.write("=" * 80 + "\n")
            
            for result in results:
                f.write(f"\n📄 ARCHIVO: {result['file']}\n")
                f.write(f"   Ruta: {result['file_path']}\n")
                
                if result.get('error'):
                    f.write(f"   ❌ ERROR: {result['error']}\n")
                    continue
                
                f.write(f"   ✅ Total de coincidencias: {result['total_matches']}\n")
                f.write(f"   📄 Páginas procesadas: {result['pages_processed']}\n")
                
                if result['matches']:
                    f.write(f"\n   COINCIDENCIAS DETALLADAS:\n")
                    f.write(f"   {'-' * 50}\n")
                    
                    # Agrupar por página
                    matches_by_page = {}
                    for match in result['matches']:
                        page = match['page']
                        if page not in matches_by_page:
                            matches_by_page[page] = []
                        matches_by_page[page].append(match)
                    
                    # Mostrar por página
                    for page in sorted(matches_by_page.keys()):
                        f.write(f"\n   📄 PÁGINA {page}:\n")
                        
                        for j, match in enumerate(matches_by_page[page], 1):
                            f.write(f"      {j}. Palabra: \"{match['matched_text']}\"\n")
                            f.write(f"         Clave buscada: {match['keyword']}\n")
                            f.write(f"         Contexto: ...{match['context']}...\n")
                            f.write(f"         {'-' * 40}\n")
                else:
                    f.write(f"   ℹ️  No se encontraron coincidencias en este archivo.\n")
                
                f.write(f"\n{'='*80}\n")
        
        # Generar reporte Excel si openpyxl está disponible
        try:
            excel_path = self.generate_excel_report(results, output_dir, timestamp, project_name)
            self.log_message(f"📊 Reporte Excel generado: {os.path.basename(excel_path)}")
        except ImportError:
            self.log_message("ℹ️  Reporte Excel no generado (openpyxl no disponible)")
        except Exception as e:
            self.log_message(f"⚠️  Error generando Excel: {str(e)}")
        
        self.log_message(f"📄 Reporte detallado guardado: {os.path.basename(report_path)}")
        return report_path

    def generate_excel_report(self, results, output_dir, timestamp, project_name):
        """Generar reporte en Excel"""
        import openpyxl
        from openpyxl.styles import Font, PatternFill, Border, Side
        
        excel_path = os.path.join(output_dir, f"reporte_{project_name}_{timestamp}.xlsx")
        wb = openpyxl.Workbook()
        
        # Hoja 1: Resumen
        ws_resumen = wb.active
        ws_resumen.title = "Resumen"
        
        # Encabezados del resumen
        headers_resumen = ['Archivo', 'Total Coincidencias', 'Páginas Procesadas', 'Estado']
        for col, header in enumerate(headers_resumen, 1):
            cell = ws_resumen.cell(row=1, column=col, value=header)
            cell.font = Font(bold=True)
            cell.fill = PatternFill(start_color="CCCCCC", end_color="CCCCCC", fill_type="solid")
        
        # Datos del resumen
        for row, result in enumerate(results, 2):
            ws_resumen.cell(row=row, column=1, value=result['file'])
            ws_resumen.cell(row=row, column=2, value=result['total_matches'])
            ws_resumen.cell(row=row, column=3, value=result['pages_processed'])
            ws_resumen.cell(row=row, column=4, value='Error' if result.get('error') else 'Procesado')
        
        # Hoja 2: Detalle de coincidencias
        ws_detalle = wb.create_sheet("Detalle de Coincidencias")
        
        headers_detalle = ['Archivo', 'Página', 'Palabra Clave', 'Texto Encontrado', 'Contexto']
        for col, header in enumerate(headers_detalle, 1):
            cell = ws_detalle.cell(row=1, column=col, value=header)
            cell.font = Font(bold=True)
            cell.fill = PatternFill(start_color="CCCCCC", end_color="CCCCCC", fill_type="solid")
        
        # Datos del detalle
        row = 2
        for result in results:
            for match in result.get('matches', []):
                ws_detalle.cell(row=row, column=1, value=result['file'])
                ws_detalle.cell(row=row, column=2, value=match['page'])
                ws_detalle.cell(row=row, column=3, value=match['keyword'])
                ws_detalle.cell(row=row, column=4, value=match['matched_text'])
                ws_detalle.cell(row=row, column=5, value=match['context'])
                row += 1
        
        # Ajustar ancho de columnas
        for ws in [ws_resumen, ws_detalle]:
            for column in ws.columns:
                max_length = 0
                column_letter = column[0].column_letter
                for cell in column:
                    try:
                        if len(str(cell.value)) > max_length:
                            max_length = len(str(cell.value))
                    except Exception:
                        pass
                adjusted_width = min(max_length + 2, 50)
                ws.column_dimensions[column_letter].width = adjusted_width
        
        wb.save(excel_path)
        return excel_path

    def show_completion_dialog(self, output_dir, report_path):
        """Mostrar diálogo de finalización con opciones"""
        dialog = tk.Toplevel(self.root)
        dialog.title("🎉 Análisis Completado")
        dialog.geometry("500x300")
        dialog.transient(self.root)
        dialog.grab_set()
        
        # Centrar el diálogo
        dialog.update_idletasks()
        x = (dialog.winfo_screenwidth() // 2) - (dialog.winfo_width() // 2)
        y = (dialog.winfo_screenheight() // 2) - (dialog.winfo_height() // 2)
        dialog.geometry(f"+{x}+{y}")
        
        main_frame = ttk.Frame(dialog, padding="20")
        main_frame.pack(fill="both", expand=True)
        
        # Título
        title_label = ttk.Label(main_frame, text="🎉 ¡Análisis Completado Exitosamente!", 
                               font=('Arial', 14, 'bold'))
        title_label.pack(pady=(0, 15))
        
        # Información
        info_text = f"""El análisis OCR se completó correctamente.
        
Los resultados se guardaron en:
📁 {os.path.abspath(output_dir)}

📄 Reporte: {os.path.basename(report_path)}

¿Qué desea hacer ahora?"""
        
        info_label = ttk.Label(main_frame, text=info_text, justify=tk.LEFT)
        info_label.pack(pady=(0, 20))
        
        # Botones
        buttons_frame = ttk.Frame(main_frame)
        buttons_frame.pack(fill="x", pady=(10, 0))
        
        def open_folder():
            # Usar la ruta absoluta capturada para asegurar la carpeta correcta
            target_dir = os.path.abspath(output_dir)
            subprocess.Popen(f'explorer "{target_dir}"')
            dialog.destroy()
        
        def open_report():
            subprocess.Popen(f'notepad "{report_path}"')
            dialog.destroy()
        
        def close_dialog():
            dialog.destroy()
        
        open_folder_btn = ttk.Button(buttons_frame, text="📁 Abrir Carpeta de Resultados", 
                                    command=open_folder, width=25)
        open_folder_btn.pack(side=tk.LEFT, padx=(0, 10))
        
        open_report_btn = ttk.Button(buttons_frame, text="📄 Abrir Reporte", 
                                    command=open_report, width=15)
        open_report_btn.pack(side=tk.LEFT, padx=(0, 10))
        
        close_btn = ttk.Button(buttons_frame, text="✅ Cerrar", 
                              command=close_dialog, width=10)
        close_btn.pack(side=tk.RIGHT)

    def analysis_finished(self):
        """Finalizar análisis y restaurar interfaz"""
        self.is_processing = False
        self.analyze_btn.config(state='normal', text="🚀 Iniciar Análisis OCR")
        self.progress_bar.stop()

    def show_help(self):
        """Mostrar ayuda"""
        help_text = """
🔍 ANALIZADOR OCR UNIVERSAL - AYUDA

Este programa analiza documentos PDF buscando palabras clave específicas.

📋 PASOS PARA USAR:

1. CONFIGURAR PROYECTO:
   - Ingrese un nombre para su proyecto
   - Opcional: agregue una descripción
   - Seleccione carpeta donde guardar resultados

2. SELECCIONAR ARCHIVOS:
   - Haga clic en "Seleccionar PDFs"
   - Elija los archivos PDF a analizar
   - Puede agregar o quitar archivos según necesite

3. CONFIGURAR PALABRAS CLAVE:
   - Agregue las palabras que desea buscar manualmente
   - CARGAR: Use archivos .txt (una palabra por línea)
   - GUARDAR: Se guardan como .txt por defecto
   - También soporta archivos .json (opcional)

4. PROCESAR:
   - Configure las opciones de análisis
   - Haga clic en "Iniciar Análisis OCR"
   - Vea el progreso en tiempo real

🎯 RESULTADOS:
- Reporte detallado con páginas y contexto
- Archivo Excel con coincidencias organizadas
- Al finalizar: botones para abrir resultados

📁 FORMATO DE ARCHIVOS TXT:
- Una palabra clave por línea
- Líneas que empiecen con # son comentarios (se ignoran)
- Ejemplo: "administración", "gobierno", "municipal"

💡 CONSEJOS:
- Use palabras clave específicas para mejores resultados
- El OCR funciona mejor con documentos de buena calidad
- Guarde sus listas de palabras para reutilizar en proyectos
- Use las plantillas incluidas como punto de partida
        """
        
        help_window = tk.Toplevel(self.root)
        help_window.title("❓ Ayuda - Analizador OCR Universal")
        help_window.geometry("600x500")
        help_window.transient(self.root)
        help_window.grab_set()
        
        help_frame = ttk.Frame(help_window, padding="20")
        help_frame.pack(fill="both", expand=True)
        
        help_scroll = scrolledtext.ScrolledText(help_frame, wrap=tk.WORD, font=('Arial', 10))
        help_scroll.pack(fill="both", expand=True)
        help_scroll.insert("1.0", help_text)
        help_scroll.config(state="disabled")
        
        close_btn = ttk.Button(help_frame, text="Cerrar", command=help_window.destroy)
        close_btn.pack(pady=(10, 0))

    def _mostrar_instalacion_exitosa(self):
        """Muestra mensaje de instalación exitosa"""
        messagebox.showinfo(
            "Instalación Completa", 
            "Las dependencias se han instalado correctamente.\nReinicie la aplicación para aplicar los cambios."
        )
    
    def _mostrar_error_instalacion(self):
        """Muestra mensaje de error en instalación"""
        messagebox.showerror(
            "Error de Instalación", 
            "Error instalando dependencias.\nVerifique la conexión a internet e inténtelo de nuevo."
        )

    def verificar_dependencias_inicial(self):
        """Verifica e instala dependencias al iniciar la aplicación"""
        try:
            # Importar el instalador automático
            from instalador_automatico import InstaladorDependencias
            
            instalador = InstaladorDependencias()
            verificacion = instalador.verificacion_completa()
            
            # Si hay dependencias faltantes, ofrecer instalación
            if verificacion['python_faltantes']:
                respuesta = messagebox.askyesno(
                    "Dependencias Requeridas",
                    f"Se detectaron {len(verificacion['python_faltantes'])} paquetes Python faltantes:\n\n" +
                    "\n".join([f"• {pkg}" for pkg in verificacion['python_faltantes']]) +
                    "\n\n¿Desea instalarlos automáticamente?",
                    icon='warning'
                )
                
                if respuesta:
                    # Instalar en segundo plano
                    def instalar_async():
                        try:
                            instalador.instalar_packages_python(verificacion['python_faltantes'])
                            self.root.after(0, self._mostrar_instalacion_exitosa)
                        except Exception:
                            self.root.after(0, self._mostrar_error_instalacion)
                    
                    threading.Thread(target=instalar_async, daemon=True).start()
                else:
                    messagebox.showwarning(
                        "Funcionalidad Limitada",
                        "Algunas funciones pueden no estar disponibles sin las dependencias requeridas."
                    )
            
            # Verificar Tesseract y archivos de idioma
            if not verificacion['tesseract_ok']:
                messagebox.showwarning(
                    "Tesseract OCR No Encontrado",
                    "Tesseract OCR no está instalado o no se encuentra.\n\n" +
                    "Para instalar Tesseract:\n" +
                    "1. Descarga desde: https://github.com/tesseract-ocr/tesseract\n" +
                    "2. Instala con configuración por defecto\n\n" +
                    "La funcionalidad OCR estará limitada sin Tesseract."
                )
            
            elif not verificacion['tessdata_info']['spa']:
                messagebox.showwarning(
                    "Archivo de Idioma Español Faltante",
                    "El archivo de idioma español (spa.traineddata) no se encuentra.\n\n" +
                    "Se usará inglés como alternativa, pero la precisión puede verse afectada."
                )
        
        except ImportError:
            # Si no se puede importar el instalador, continuar normalmente
            pass
        except Exception as e:
            messagebox.showwarning(
                "Error de Verificación",
                f"Error verificando dependencias:\n{str(e)}\n\nLa aplicación continuará normalmente."
            )

def main():
    root = tk.Tk()
    app = UniversalOCRAnalyzer(root)
    
    # Centrar ventana
    root.update_idletasks()
    x = (root.winfo_screenwidth() // 2) - (root.winfo_width() // 2)
    y = (root.winfo_screenheight() // 2) - (root.winfo_height() // 2)
    root.geometry(f"+{x}+{y}")
    
    root.mainloop()

if __name__ == "__main__":
    main()
